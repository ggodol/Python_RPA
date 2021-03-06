"""
6. 파일 열기
https://youtu.be/exgO1LFl9x8?t=2077
"""
from openpyxl import load_workbook # 파일 불러 오기
wb = load_workbook("sample.xlsx") # sample.xlsx 파일에서 wb 을 불러옴
ws = wb.active # 활성화된 sheet

# cell 데이터 불러오기
# for x in range(1,11):
#     for y in range(1,11):
#         print(ws.cell(row=x, column=y).value, end=" ") # 1 2 3 4
#     print()

# cell 갯수를 모를 때
for x in range(1,ws.max_row + 1):
    for y in range(1,ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=" ") # 1 2 3 4
    print()